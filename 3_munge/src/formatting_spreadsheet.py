# -*- coding: utf-8 -*-
"""
Created on Fri Mar 29 10:53:47 2024

@author: AbnerBogan
"""

import openpyxl 
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, PatternFill, Side


def load_spreadsheet(tmp_file):
    """
    Load spreadsheet from file.

    Parameters
    ----------
    tmp_file : string
        Name of temporary file containing spreadsheet of cluster dataset metadata. 

    Returns
    -------
    wb : openpxyl workbook
        Workbook object containing spreadsheet where each sheet contains dataset metadata of a given cluster.

    """
 
    wb = openpyxl.load_workbook(tmp_file)
        
    return wb


def get_max_len_column_headers(ws):
    """
    Gets the maximum length of all the column headers in worksheet based 
    on number of characters.

    Parameters
    ----------
    ws : openpxyl worksheet
        Worksheet object containing dataset metadata of a given cluster.

    Returns
    -------
    max_length : int
        Maximum length of column headers in worksheet.

    """
    
    # start loop for iterating through column headers
    for i, col in enumerate(ws.columns):

        # get length of column header
        length = len(col[0].value)
        
        # update max_length if length is larger
        if i == 0:
            max_length = length
        elif length > max_length:
            max_length = length
               
    return max_length

def adjust_column_width(ws,max_length,date_format ='YYYY-MM-DD HH:MM:SS'):
    """
    Adjust column width in worksheet based on maximum length of column headers, 
    along with standardized datetime format.

    Parameters
    ----------
    ws : openpxyl worksheet
        Worksheet object containing dataset metadata of a given cluster.
    max_length : int
        Maximum length of column headers in worksheet.
    date_format : string, optional
        Standardized datetime format used in spreadsheet. The default is 'YYYY-MM-DD HH:MM:SS'.

    Returns
    -------
    None.

    """
    
    # iterate over columns in spreadsheet
    for i, col in enumerate(ws.columns):
    
        # retrieve the associated column letter
        col_letter = get_column_letter(i+1)

        # set column width; for date columns, the column needs to be wider 
        # so the date is visible and doesn't appear as '#########'
        if col[0].value == 'dateCreated' or col[0].value == 'datePublished':    
            ws.column_dimensions[col_letter].width = len(date_format)
        # set column width for other columns based on max length
        else:
            ws.column_dimensions[col_letter].width = max_length 
        
def set_horizontal_center_header_row(ws):
    """
    Set horizontal center to values in header row of spreadsheet.

    Parameters
    ----------
    ws : openpxyl worksheet
        Worksheet object containing dataset metadata of a given cluster.

    Returns
    -------
    None.

    """
    
    # create loop to iterate over cells in first row across all columns
    for col in ws.iter_cols(min_row=1, max_row=1,max_col=ws.max_column):
        for cell in col:
            # set horizontal center alignment to cell
            cell.alignment = Alignment(horizontal="center")    
            
def make_borders(ws):
    """
    Make horizontal border under first row across all column headers

    Parameters
    ----------
    ws : openpxyl worksheet
        Worksheet object containing dataset metadata of a given cluster.

    Returns
    -------
    None.

    """
    
    # define thick bottom border style
    thick_bottom_border = Border(bottom=Side(style='thick'))
    
    # create loop to iterate over cells in the first row across all columns 
    for col in ws.iter_cols(min_row=1, max_row=1, max_col=ws.max_column):
        for cell in col:
            # apply the thick bottom border to cell
            cell.border = thick_bottom_border

def color_published_rows(ws,color='00FF00'):
    """
    Highlight rows corresponding to published datasets

    Parameters
    ----------
    ws : openpxyl worksheet
        Worksheet object containing dataset metadata of a given cluster.
    color : string, optional
        Hex color code. The default is '00FF00' (green).

    Returns
    -------
    None.

    """
    
    # define color fill for highlighting cells
    color_fill = PatternFill(start_color=color,end_color=color,fill_type='solid')

    num_rows = ws.max_row
    num_cols = ws.max_column

    # iterate over rows to search for published datasets
    for row in ws.iter_rows(min_row=2, max_row=num_rows,max_col=num_cols):
        # check if cell in last column (access) indicates published dataset
        if row[num_cols-1].value == 'PUBLISHED':
            # fill cells in row with color fill
            for cell in row:
                cell.fill = color_fill
                

def save_spreadsheet(wb,out_file):
    """
    Save styled spreadsheet to file.

    Parameters
    ----------
    wb : openpxyl workbook
        Workbook object containing styled spreadsheet where each sheet contains dataset metadata of a given cluster.
    out_file : string
        Name of final output file containing master spreadsheet.

    Returns
    -------
    None.

    """
    
    
    wb.save(out_file)
                

def main(tmp_file,out_file):
    """
    Main function.

    Parameters
    ----------
    tmp_file : string
        Name of temporary file containing spreadsheet of cluster dataset metadata.
    out_file : string
        Name of final output file containing master spreadsheet.

    Returns
    -------
    None.

    """
    
    # load spreadsheet from file as openpyxl workbook object
    wb = load_spreadsheet(tmp_file)
    
    # iterate over sheets in workbook
    for ws in wb:
        
        # get max length of column headers by number of characters
        max_length = get_max_len_column_headers(ws)
        
        # use this max length to adjust the column width to improve readability
        adjust_column_width(ws, max_length)
        
        # set horizontal center to cells in header row
        set_horizontal_center_header_row(ws)
        
        # make border under header row
        make_borders(ws)
        
        # highlight/color rows corresponding to datasets with published sharing status
        color_published_rows(ws)
    
    # save final output spreadsheet to file
    save_spreadsheet(wb, out_file)
    
    
    
if __name__ == '__main__':
    
    input_filename = snakemake.input['in_filename']

    output_filename = snakemake.output['out_filename']
    
    main(input_filename,output_filename)
    
    